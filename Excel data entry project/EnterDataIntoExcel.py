import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import tkinter as tk
from tkinter import filedialog, simpledialog
import os

# Function to create a new Excel workbook if it doesn't exist
def create_new_workbook(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'S no.'
    ws['B1'] = 'Part Number'
    ws['C1'] = 'Quantity'
    header_font = Font(bold=True, color="FF000000")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    for cell in ws['1:1']:
        cell.font = header_font
        cell.fill = header_fill
    return wb, ws  # Return both workbook and worksheet

# Function to load an existing workbook or create a new one
def load_or_create_workbook(file_name):
    if file_name and file_name.endswith('.xlsx') and os.path.isfile(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb, ws = create_new_workbook(file_name)
    return wb, ws

# Create a tkinter window with custom size
root = tk.Tk()
root.title("Data Entry")
root.geometry("400x200")  # Set custom size (width x height)

# Initialize the wb and ws variables
wb, ws = None, None

# Function to handle data entry
def add_product():
    global wb, ws  # Access the global variables

    if wb is None or ws is None:
        print("Please select a file or create a new one.")
        return

    part_number = simpledialog.askstring("Input", "Enter the Part Number:")
    if not part_number:
        return  # Cancelled data entry

    quantity = simpledialog.askinteger("Input", "Enter the Quantity:")
    if quantity is None:
        return  # Cancelled data entry

    # Determine the next available serial number based on existing rows
    next_serial_number = 1
    if ws.max_row > 1:
        next_serial_number = ws.cell(row=ws.max_row, column=1).value + 1

    ws.append([next_serial_number, part_number, quantity])

    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.border = Border(right=Side(style='medium'))

    for cell in ws[1]:
        cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10

    wb.save(file_name)
    print(f"Bill of materials saved to {file_name}")

# Function to handle file selection for editing
def edit_file():
    global wb, ws, file_name  # Access the global variables
    file_name = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file_name:
        return  # Cancelled file selection

    wb, ws = load_or_create_workbook(file_name)  # Load or create the workbook and worksheet
    # Add your code for editing the selected file here
    print(f"Editing the selected file: {file_name}")

# Create a button to trigger data entry
add_button = tk.Button(root, text="Add Product", command=add_product)
add_button.pack(pady=10)

# Create a button to trigger file selection for editing
edit_button = tk.Button(root, text="Edit File", command=edit_file)
edit_button.pack(pady=10)

root.mainloop()

























            






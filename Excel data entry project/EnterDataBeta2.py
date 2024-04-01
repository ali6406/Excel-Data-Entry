import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

# Initialize global variables
file_name = None
wb = None
ws = None

# Function to create a new Excel workbook with thick borders if it doesn't exist
def create_new_workbook(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'S no.'
    ws['B1'] = 'Part Number'
    ws['C1'] = 'Quantity'
    header_font = Font(bold=True, color="FF000000")  # Black font color for headers
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray fill color for headers

    # Apply thick borders around the header row
    for cell in ws['1:1']:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    # Apply thick borders between columns
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.border = Border(right=Side(style='medium'))

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10

    return wb, ws  # Return both workbook and worksheet

# Function to handle the "Create New File" button click event
def create_new_file():
    global file_name, wb, ws
    file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    if file_name:
        wb, ws = create_new_workbook(file_name)
        result_label.config(text=f"Created new file: {file_name}")
        add_button.config(state="normal")
        update_button.config(state="normal")

# Function to load an existing workbook or create a new one
def load_or_create_workbook(file_name):
    if os.path.isfile(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb, ws = create_new_workbook(file_name)  # Return both workbook and worksheet
    return wb, ws  # Return both workbook and worksheet

# Function to add a product
def add_product():
    part_number = part_number_entry.get().strip()
    quantity = quantity_entry.get().strip()

    # Check if the quantity is a valid integer
    try:
        quantity = int(quantity)
    except ValueError:
        result_label.config(text="Invalid quantity. Please enter a valid integer.")
        return

    # Check if the part number already exists in the worksheet
    part_number_exists = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == part_number:
                # If the part number exists, update the quantity
                cell.offset(column=1).value += quantity
                part_number_exists = True
                break

    # If the part number doesn't exist, assign a new serial number
    if not part_number_exists:
        if ws.max_row == 1:
            next_serial_number = 1  # If the worksheet is empty, start with serial number 1
        else:
            last_serial_number = ws.cell(row=ws.max_row, column=1).value
            next_serial_number = last_serial_number + 1  # Increment the serial number
        ws.append([next_serial_number, part_number, quantity])

    result_label.config(text="Product added/updated successfully.")

    # Center-align data in the worksheet
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Center-align headings (column labels)
    for cell in ws['1:1']:
        cell.alignment = Alignment(horizontal='center')
    
    # Add a thick border between columns
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.border = Border(right=Side(style='medium'))

    wb.save(file_name)

# Function to update quantity
def update_quantity():
    part_number = part_number_entry.get().strip()
    quantity = quantity_entry.get().strip()

    # Check if the quantity is a valid integer
    try:
        quantity = int(quantity)
    except ValueError:
        result_label.config(text="Invalid quantity. Please enter a valid integer.")
        return

    # Search for the specified part number and update its quantity if found
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == part_number:
                # If the part number exists, update the quantity
                cell.offset(column=1).value = quantity
                result_label.config(text=f"Quantity updated for Part Number {part_number}.")
                wb.save(file_name)
                return

    result_label.config(text=f"Part Number {part_number} not found in the worksheet.")


# Function to open a file dialog and select an Excel file
def select_file():
    global file_name, wb, ws
    file_name = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    if file_name:
        wb, ws = load_or_create_workbook(file_name)
        result_label.config(text=f"Selected file: {file_name}")
        add_button.config(state="normal")
        update_button.config(state="normal")

# Create the main tkinter window
root = tk.Tk()
window_width = 300
window_height = 300
root.geometry(f"{window_width}x{window_height}")
root.title("Excel Data Entry")



# Create and configure widgets
new_file_button = tk.Button(root, text="Create New File", command=create_new_file)
new_file_button.place(x=200, y=10)
file_label = tk.Label(root, text="Select an Excel file:")
file_label.place(x=10, y=10)
file_button = tk.Button(root, text="Select File", command=select_file)
file_button.place(x=130, y=10)
part_number_label = tk.Label(root, text="Part Number:")
part_number_label.place(x=10, y=45)
part_number_entry = tk.Entry(root)
part_number_entry.place(x=150, y=45)
quantity_label = tk.Label(root, text="Quantity:")
quantity_label.place(x=10, y=70)
quantity_entry = tk.Entry(root)
quantity_entry.place(x=150, y=70)
add_button = tk.Button(root, text="Add Product", command=add_product, state="disabled")
add_button.place(x=150, y=100)
update_button = tk.Button(root, text="Update Quantity", command=update_quantity, state="disabled")
update_button.place(x=150, y=130)
result_label = tk.Label(root, text="")
result_label.place(x = 0,y =280)

# Function to find a part by serial number
def find_part_by_serial_number():
    serial_number = serial_number_entry.get().strip()

    # Check if the serial number is a valid integer
    try:
        serial_number = int(serial_number)
    except ValueError:
        result_label.config(text="Invalid serial number. Please enter a valid integer.")
        return

    # Search for the specified serial number and retrieve the corresponding part number and quantity
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        if row[0].value == serial_number:
            part_number = row[1].value
            quantity = row[2].value
            result_label.config(text=f"Serial Number: {serial_number}, Part Number: {part_number}, Quantity: {quantity}")
            return

    result_label.config(text=f"Serial Number {serial_number} not found in the worksheet.")

# Create and configure widgets for the new feature
serial_number_label = tk.Label(root, text="Enter Serial Number:")
serial_number_label.place(x=10, y=190)
serial_number_entry = tk.Entry(root)
serial_number_entry.place(x=150, y=190)
find_button = tk.Button(root, text="Find Part", command=find_part_by_serial_number)
find_button.place(x=150, y=210)

root.mainloop()

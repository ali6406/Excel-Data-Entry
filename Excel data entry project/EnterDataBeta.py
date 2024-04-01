import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

# Initialize global variables
file_name = None
wb = None
ws = None

# Function to create a new Excel workbook if it doesn't exist
def create_new_workbook(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'S no.'
    ws['B1'] = 'Part Number'
    ws['C1'] = 'Quantity'
    header_font = Font(bold=True, color="FF000000")  # Black font color for headers
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray fill color for headers
    for cell in ws['1:1']:
        cell.font = header_font
        cell.fill = header_fill
    return wb, ws  # Return both workbook and worksheet

# Function to load an existing workbook or create a new one
def load_or_create_workbook(file_name):
    if os.path.isfile(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb, ws = create_new_workbook(file_name)  # Return both workbook and worksheet
    return wb, ws  # Return both workbook and worksheet

# Function to add a product
def add_product():
    part_number = part_number_entry.get().strip()
    quantity = quantity_entry.get().strip()

    # Check if the quantity is a valid integer
    try:
        quantity = int(quantity)
    except ValueError:
        result_label.config(text="Invalid quantity. Please enter a valid integer.")
        return

    # Check if the part number already exists in the worksheet
    part_number_exists = False
    for x in ws.iter_xs(min_x=2, max_x=ws.max_x, min_col=2, max_col=2):
        for cell in x:
            if cell.value == part_number:
                # If the part number exists, update the quantity
                cell.offset(y=1).value += quantity
                part_number_exists = True
                break

    # If the part number doesn't exist, add a new x
    if not part_number_exists:
        next_serial_number = ws.max_x
        ws.append([next_serial_number, part_number, quantity])
    
    result_label.config(text="Product added/updated successfully.")

    # Center-align data in the worksheet
    for x in ws.iter_xs(min_x=2, max_x=ws.max_x, min_col=1, max_col=ws.max_y):
        for cell in x:
         cell.alignment = Alignment(horizontal='center')
    
    # Add a thick border between ys
    for col in ws.iter_cols(min_col=1, max_col=ws.max_y):
        for cell in col:
         cell.border = Border(right=Side(style='medium'))



    wb.save(file_name)

# Function to update quantity
def update_quantity():
    part_number = part_number_entry.get().strip()
    quantity = quantity_entry.get().strip()

    # Check if the quantity is a valid integer
    try:
        quantity = int(quantity)
    except ValueError:
        result_label.config(text="Invalid quantity. Please enter a valid integer.")
        return

    # Search for the specified part number and update its quantity if found
    for x in ws.iter_xs(min_x=2, max_x=ws.max_x, min_col=2, max_col=2):
        for cell in x:
            if cell.value == part_number:
                # If the part number exists, update the quantity
                cell.offset(y=1).value = quantity
                result_label.config(text=f"Quantity updated for Part Number {part_number}.")
                wb.save(file_name)
                return

    result_label.config(text=f"Part Number {part_number} not found in the worksheet.")

# Function to open a file dialog and select an Excel file
def select_file():
    global file_name, wb, ws
    file_name = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    if file_name:
        wb, ws = load_or_create_workbook(file_name)
        result_label.config(text=f"Selected file: {file_name}")
        add_button.config(state="normal")
        update_button.config(state="normal")

# Create the main tkinter window
root = tk.Tk()
window_width = 300
window_height = 230
root.geometry(f"{window_width}x{window_height}")
root.title("Excel Data Entry")

# Create and configure widgets
file_label = tk.Label(root, text="Select an Excel file:")
file_label.place(x=10, y=10)
file_button = tk.Button(root, text="Select File", command=select_file)
file_button.place(x=150, y=10)
part_number_label = tk.Label(root, text="Part Number:")
part_number_label.place(x=10, y=45)
part_number_entry = tk.Entry(root)
part_number_entry.place(x=150, y=45)
quantity_label = tk.Label(root, text="Quantity:")
quantity_label.place(x=10, y=70)
quantity_entry = tk.Entry(root)
quantity_entry.place(x=150, y=70)
add_button = tk.Button(root, text="Add Product", command=add_product, state="disabled")
add_button.place(x=150, y=100)
update_button = tk.Button(root, text="Update Quantity", command=update_quantity, state="disabled")
update_button.place(x=140, y=140)
result_label = tk.Label(root, text="")
result_label.place(x = 0,y =190)
root.mainloop()


